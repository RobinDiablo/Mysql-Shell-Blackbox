import psutil
import re
from datetime import timedelta,datetime
from openpyxl import load_workbook
import csv
import os
import tkinter as tk


def extractor():

            dateTimeObj=datetime.utcnow()
            cal={"Jan":"01","Feb":"02","Mar":"03","Apr":"04","May":"05","Jun":"06","Jul":"07","Aug":"08","Sep":"09","Oct":"10","Nov":"11","Dec":"12"}
            m=dateTimeObj.strftime("%b")
            for x, y in cal.items():
              if m==x:
                m=y
            time=dateTimeObj.strftime("%Y-")+str(m)+dateTimeObj.strftime("-%dT%H:%M:%S.%f")

            wb = load_workbook(r'C:\Users\Balaji Vyshnavy\Desktop\state2.xlsx')
            sheet = wb.active
              

            badw = ['GRANT PROXY ON','root@localhost on mysql using TCP/IP' ,'@is_mysql_encryopted' ,'@@' ,'DROP PREPARE stmt','SHOW FULL COLUMNS FROM','show columns from',' EXECUTE stmt','PREPARE stmt FROM @str','   DROP PREPARE stmt','   DROP PREPARE stmt ',' create table if not exsts  ','rollback','CREATE TABLE IF NOT EXISTS',' oracle ','@cmd','SHOW GLOBAL STATUS','GRANT SELECT ON']
            with open('C:/ProgramData/MySQL/MySQL Server 8.0/Data/LAPTOP-67H7KN81.log', 'r') as oldfile, open('C:/Users/Balaji Vyshnavy/Desktop/logs/temp.txt', 'w+') as newfile:
                for line in oldfile:
                    line.lstrip()
                    if not any(bw in line for bw in badw) and line[0].isdigit():
                            newfile.write(line)

            oldfile.close()
            newfile.close()


            def search_index(time,data,l):
              i=0
              while i<l:
                t=res[i]
                t=t[0:16]
                i=i+1
                if t==time:
                  return i
              return 0


            trunc=open('C:/Users/Balaji Vyshnavy/Desktop/logs/evidence.csv', 'a')
            trunc.truncate()
            trunc.close()
            temp=open('C:/Users/Balaji Vyshnavy/Desktop/logs/temp.txt', 'r+')
            res=temp.readlines()
            l=len(res)
            t=res[l-1]
            t=t[0:16]
            print(t)
            s=sheet['B1'].value
            s=s[0:16]
            print(s)
            cursor=search_index(s,res,l)
            save=open('C:/Users/Balaji Vyshnavy/Desktop/logs/evidence.csv', 'a')
            print("Total new records:",l-cursor)
            while cursor<l:
              save.write(res[cursor])
              cursor=cursor+1



            if os.path.isfile('C:/Users/Balaji Vyshnavy/AppData/Roaming/MySQL/mysqlsh/history.sql')==True and  os.path.isfile('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.sql')==False:
              shutil.move('C:/Users/Balaji Vyshnavy/AppData/Roaming/MySQL/mysqlsh/history.sql','C:/Users/Balaji Vyshnavy/Desktop/logs/temp/')
            if os.path.isfile('C:/Users/Balaji Vyshnavy/AppData/Roaming/MySQL/mysqlsh/history.js')==True and os.path.isfile('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.js')==False:
              shutil.move('C:/Users/Balaji Vyshnavy/AppData/Roaming/MySQL/mysqlsh/history.js','C:/Users/Balaji Vyshnavy/Desktop/logs/temp/')
            if os.path.isfile('C:/Users/Balaji Vyshnavy/AppData/Roaming/MySQL/mysqlsh/history.py')==True and os.path.isfile('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.py')==False:
              shutil.move('C:/Users/Balaji Vyshnavy/AppData/Roaming/MySQL/mysqlsh/history.py','C:/Users/Balaji Vyshnavy/Desktop/logs/temp/')

            def append(file1,file2,time):
              with open(file1, 'r') as oldfile, open(file2, 'w') as newfile:
                newfile.write(time)
                for line in oldfile:
                    line.lstrip()
                    newfile.write(line)


            append('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.sql','C:/Users/Balaji Vyshnavy/Desktop/logs/locker/history.sql',time)
            append('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.py','C:/Users/Balaji Vyshnavy/Desktop/logs/locker/history.py',time)
            append('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.js','C:/Users/Balaji Vyshnavy/Desktop/logs/locker/history.js',time)

            def uploadevidence(file1,file2):
                with open(file1, 'r') as oldfile, open(file2, 'a') as newfile:
                  for line in oldfile:
                      line.lstrip()
                      newfile.write(line)

            uploadevidence('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.sql','C:/Users/Balaji Vyshnavy/Desktop/logs/evidence.csv')  
            uploadevidence('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.js','C:/Users/Balaji Vyshnavy/Desktop/logs/evidence.csv')
            uploadevidence('C:/Users/Balaji Vyshnavy/Desktop/logs/temp/history.py','C:/Users/Balaji Vyshnavy/Desktop/logs/evidence.csv')  


wb = load_workbook(r'C:\Users\Balaji Vyshnavy\Desktop\state.xlsx')
sheet = wb.active

class ExampleApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.geometry('200x90')
        self.label = tk.Label(self, text="",font=("Courier", 50), width=300,height=200,bd=10,bg='sky blue')
        self.label.place(height=90,width=200)
        self.remaining = 0
        self.countdown(10)

    def countdown(self, remaining = None):
        if remaining is not None:
            self.remaining = remaining

        if self.remaining <= 0:
            self.label.configure(text="Time's up!",font=("Courier", 25),width=300,height=200,bd=10,bg='sky blue')
        else:
            self.label.configure(text="%d" % self.remaining,width=300,height=200,bd=10,bg='sky blue')
            self.remaining = self.remaining - 1
            self.after(1000, self.countdown)



dateTimeObj = datetime.now()
timenow= dateTimeObj.strftime("%H:%M:%S")
x,y,z=timenow.split(':')
x=int(x)
y=int(y)
z=int(z)
t1 = timedelta(hours=x, minutes=y,seconds=z)
t2 = timedelta(hours=0, minutes=20,seconds=00)

def find():
    for p in psutil.process_iter():
        line=p.name
        pid=p.pid
        line = str(line)
        line=line[49:-2]
        bad_chars = [", name='", "', started='"]
        bc=["'"]
        test_string = line
        test_string = str(test_string)
        for i in bad_chars:
            test_string = test_string.replace(i, ';')
        for i in bc:
            test_string = test_string.replace(i, '')
        time=test_string[-8:]
        y= test_string.rpartition(';')
        y=y[0]
        name = "".join(re.split("[^a-zA-Z]*", y))
        name=str(name)
        msg='mysqlshexe'
        if name==msg :
            return 1,time,pid
            break
    return 0,0,0

def session(time):
    print("running",pid)
    time=str(time)
    print(time)
    l,m,n=time.split(':')
    l = int(l)
    m = int(m)
    n = int(n)
    t3 = timedelta(hours=l, minutes=m, seconds=n)
    print(t1)
    print(t3)
    t4=t1-t3
    t4=str(t4)
    print(t4)
    p,q,r=t4.split(':')
    p=int(p)
    q=int(q)
    if q>=1 or p>=1:
        return 1
    else:
        return 0    

x,time,pid=find()
if x==1:
    y=session(time)

if x==1 and y==1:
    print("killing")
    print(pid)
    if __name__ == "__main__":
        app = ExampleApp()
        app.after(12000, lambda: app.destroy())
        app.mainloop()
    os.kill(pid,pid)
    print("calling extractor and AI engine\nsaving state")
    extractor()
    sheet['A1'] = 'saved'
    sheet['B1'] = timenow
    wb.save(r'C:\Users\Balaji Vyshnavy\Desktop\state.xlsx')
    print("killed")
    #statefile.close()

elif x==1 and y!=1:
    sheet['A1'] ='unsaved'
    wb.save(r'C:\Users\Balaji Vyshnavy\Desktop\state.xlsx')

elif x==0 and sheet['A1'].value=='unsaved':
    print("saving")
    sheet['A1'] = 'saved'
    sheet['B1'] = timenow
    extractor()
    wb.save(r'C:\Users\Balaji Vyshnavy\Desktop\state.xlsx')











